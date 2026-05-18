import telebot
from telebot import types
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
from bot.config import ADMIN_ID
from bot.database.db import (
    get_cities,
    add_city,
    delete_city,
    get_subjects,
    add_subject,
    delete_subject,
    add_teacher,
    delete_teacher,
    add_center,
    delete_center,
    add_university,
    delete_university,
    get_stats,
    get_all_user_ids,
    save_broadcast,
    get_last_broadcast,
    get_teachers,
    get_centers,
    get_universities,
    get_positions_list,
    set_item_position,
    get_views_list,
)
from bot.utils.keyboards import (
    admin_main_keyboard,
    admin_cities_keyboard,
    admin_subjects_keyboard,
    skip_media_keyboard,
    skip_button_keyboard,
    confirm_broadcast_keyboard,
    admin_select_keyboard,
    broadcast_btn_step_keyboard,
    position_select_keyboard,
    edit_positions_menu_keyboard,
    items_list_keyboard,
    move_item_keyboard,
    stats_views_keyboard,
)

admin_state = {}


def is_admin(user_id):
    return user_id == ADMIN_ID


def register(bot: telebot.TeleBot):
    # ─── /admin ──────────────────────────────────────────────────

    @bot.message_handler(commands=["admin"])
    def admin_panel(message):
        if not is_admin(message.from_user.id):
            bot.send_message(
                message.chat.id, "❌ <b>Ruxsat yo'q!</b>", parse_mode="HTML"
            )
            return
        admin_state.pop(message.from_user.id, None)
        bot.send_message(
            message.chat.id,
            "⚙️ <b>Admin Panel</b>\n\n👇 Amalni tanlang:",
            parse_mode="HTML",
            reply_markup=admin_main_keyboard(),
        )

    # ─── ORQAGA ──────────────────────────────────────────────────

    @bot.callback_query_handler(func=lambda c: c.data == "adm:back")
    def adm_back(call):
        if not is_admin(call.from_user.id):
            return
        admin_state.pop(call.from_user.id, None)
        try:
            bot.edit_message_text(
                "⚙️ <b>Admin Panel</b>\n\n👇 Amalni tanlang:",
                call.message.chat.id,
                call.message.message_id,
                parse_mode="HTML",
                reply_markup=admin_main_keyboard(),
            )
        except Exception:
            bot.send_message(
                call.message.chat.id,
                "⚙️ <b>Admin Panel</b>",
                parse_mode="HTML",
                reply_markup=admin_main_keyboard(),
            )
        bot.answer_callback_query(call.id)

    # ─── STATISTIKA ──────────────────────────────────────────────

    @bot.callback_query_handler(func=lambda c: c.data == "adm:stats")
    def adm_stats(call):
        if not is_admin(call.from_user.id):
            return
        total, active = get_stats()
        bot.edit_message_text(
            f"📊 <b>Statistika</b>\n\n"
            f"👥 Jami foydalanuvchilar: <b>{total}</b>\n"
            f"🟢 Faol (7 kun): <b>{active}</b>\n"
            f"🔴 Nofaol: <b>{total - active}</b>\n\n"
            f"📈 Ko'rishlarni ko'rish uchun tanlang:",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="HTML",
            reply_markup=stats_views_keyboard(),
        )
        bot.answer_callback_query(call.id)

    @bot.callback_query_handler(func=lambda c: c.data.startswith("adm:stats_views:"))
    def adm_stats_views(call):
        if not is_admin(call.from_user.id):
            return
        table = call.data.split(":")[2]
        labels = {
            "teachers": ("👨‍🏫 Ustozlar", "ustoz"),
            "centers": ("🏫 Markazlar", "markaz"),
            "universities": ("🎓 Universitetlar", "universitet"),
        }
        title, _ = labels.get(table, ("Ko'rishlar", ""))
        items = get_views_list(table)
        if not items:
            bot.answer_callback_query(call.id, "⚠️ Hozircha ma'lumot yo'q!", show_alert=True)
            return

        kb = types.InlineKeyboardMarkup(row_width=1)
        for i, item in enumerate(items, 1):
            views = item.get("views", 0)
            views_str = f"{views:,}".replace(",", " ")
            name = (item.get("name") or "")[:30]
            kb.add(types.InlineKeyboardButton(
                f"{i}. {name} — {views_str} 👁",
                callback_data="noop"
            ))
        kb.add(types.InlineKeyboardButton("🔙 Orqaga", callback_data="adm:stats"))

        try:
            bot.edit_message_reply_markup(
                call.message.chat.id,
                call.message.message_id,
                reply_markup=kb,
            )
        except Exception:
            pass
        bot.answer_callback_query(call.id)

    # ─── EXCEL EXPORT ────────────────────────────────────────────

    @bot.callback_query_handler(func=lambda c: c.data == "adm:xlsx_teachers")
    def adm_xlsx_teachers(call):
        if not is_admin(call.from_user.id):
            return
        rows = get_teachers()
        _send_xlsx(
            bot, call,
            title="Ustozlar",
            headers=["ID", "Ism", "Shahar", "Fan", "Tur", "Tavsif", "Qo'shilgan"],
            data=[[r["id"], r["name"], r["city"], r["subject"], r["teach_type"],
                   r.get("description", ""),
                   r.get("created_at", "")] for r in rows],
            filename="ustozlar.xlsx"
        )

    @bot.callback_query_handler(func=lambda c: c.data == "adm:xlsx_centers")
    def adm_xlsx_centers(call):
        if not is_admin(call.from_user.id):
            return
        rows = get_centers()
        _send_xlsx(
            bot, call,
            title="O'quv markazlari",
            headers=["ID", "Nomi", "Shahar", "Tavsif",
                     "Tugma matni", "URL", "Qo'shilgan"],
            data=[[r["id"], r["name"], r["city"], r.get("description", ""),
                   r.get("button_text", ""), r.get("button_url", ""),
                   r.get("created_at", "")] for r in rows],
            filename="markazlar.xlsx"
        )

    @bot.callback_query_handler(func=lambda c: c.data == "adm:xlsx_unis")
    def adm_xlsx_unis(call):
        if not is_admin(call.from_user.id):
            return
        rows = get_universities()
        _send_xlsx(
            bot, call,
            title="Universitetlar",
            headers=["ID", "Nomi", "Tur", "Tavsif",
                     "Tugma matni", "URL", "Qo'shilgan"],
            data=[[r["id"], r["name"], r["uni_type"], r.get("description", ""),
                   r.get("button_text", ""), r.get("button_url", ""),
                   r.get("created_at", "")] for r in rows],
            filename="universitetlar.xlsx"
        )

    # ─── SHAHARLAR ───────────────────────────────────────────────

    @bot.callback_query_handler(func=lambda c: c.data == "adm:cities")
    def adm_cities(call):
        if not is_admin(call.from_user.id):
            return
        cities = get_cities()
        bot.edit_message_text(
            f"🏙️ <b>Shaharlar</b> ({len(cities)} ta)\n🗑 O'chirish uchun bosing:",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="HTML",
            reply_markup=admin_cities_keyboard(cities),
        )
        bot.answer_callback_query(call.id)

    @bot.callback_query_handler(func=lambda c: c.data.startswith("adm_delcity:"))
    def adm_del_city(call):
        if not is_admin(call.from_user.id):
            return
        city = call.data.split(":", 1)[1]
        delete_city(city)
        cities = get_cities()
        bot.edit_message_text(
            f"🏙️ <b>Shaharlar</b> ({len(cities)} ta)\n🗑 O'chirish uchun bosing:",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="HTML",
            reply_markup=admin_cities_keyboard(cities),
        )
        bot.answer_callback_query(call.id, f"✅ '{city}' o'chirildi!")

    @bot.callback_query_handler(func=lambda c: c.data == "adm:add_city")
    def adm_add_city_start(call):
        if not is_admin(call.from_user.id):
            return
        admin_state[call.from_user.id] = {"action": "add_city"}
        bot.edit_message_text(
            "🏙️ <b>Yangi shahar qo'shish</b>\n\n✏️ Shahar nomini kiriting:",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="HTML",
            reply_markup=admin_main_keyboard(),
        )
        bot.answer_callback_query(call.id)

    # ─── FANLAR ──────────────────────────────────────────────────

    @bot.callback_query_handler(func=lambda c: c.data == "adm:subjects")
    def adm_subjects(call):
        if not is_admin(call.from_user.id):
            return
        subjects = get_subjects()
        bot.edit_message_text(
            f"📚 <b>Fanlar</b> ({len(subjects)} ta)\n🗑 O'chirish uchun bosing:",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="HTML",
            reply_markup=admin_subjects_keyboard(subjects),
        )
        bot.answer_callback_query(call.id)

    @bot.callback_query_handler(func=lambda c: c.data.startswith("adm_delsub:"))
    def adm_del_sub(call):
        if not is_admin(call.from_user.id):
            return
        subj = call.data.split(":", 1)[1]
        delete_subject(subj)
        subjects = get_subjects()
        bot.edit_message_text(
            f"📚 <b>Fanlar</b> ({len(subjects)} ta)\n🗑 O'chirish uchun bosing:",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="HTML",
            reply_markup=admin_subjects_keyboard(subjects),
        )
        bot.answer_callback_query(call.id, f"✅ '{subj}' o'chirildi!")

    @bot.callback_query_handler(func=lambda c: c.data == "adm:add_subject")
    def adm_add_subj_start(call):
        if not is_admin(call.from_user.id):
            return
        admin_state[call.from_user.id] = {"action": "add_subject"}
        bot.edit_message_text(
            "📚 <b>Yangi fan qo'shish</b>\n\n✏️ Fan nomini kiriting:",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="HTML",
            reply_markup=admin_main_keyboard(),
        )
        bot.answer_callback_query(call.id)

    # ─── O'QITUVCHI QO'SHISH ──────────────────────────────────────

    @bot.callback_query_handler(func=lambda c: c.data == "adm:add_teacher")
    def adm_add_teacher(call):
        if not is_admin(call.from_user.id):
            return
        admin_state[call.from_user.id] = {
            "action": "add_teacher",
            "step": "name",
            "data": {},
        }
        bot.edit_message_text(
            "👨‍🏫 <b>Ustoz qo'shish</b>\n\n📝 <b>1-qadam:</b> To'liq ismini kiriting:",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="HTML",
            reply_markup=admin_main_keyboard(),
        )
        bot.answer_callback_query(call.id)

    @bot.callback_query_handler(func=lambda c: c.data == "adm:del_teacher")
    def adm_del_teacher_start(call):
        if not is_admin(call.from_user.id):
            return
        admin_state[call.from_user.id] = {"action": "del_teacher"}
        bot.edit_message_text(
            "🗑️ <b>Ustoz o'chirish</b>\n\n🔢 Ustoz <b>ID</b> sini kiriting:",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="HTML",
            reply_markup=admin_main_keyboard(),
        )
        bot.answer_callback_query(call.id)

    # ─── MARKAZ ──────────────────────────────────────────────────

    @bot.callback_query_handler(func=lambda c: c.data == "adm:add_center")
    def adm_add_center(call):
        if not is_admin(call.from_user.id):
            return
        admin_state[call.from_user.id] = {
            "action": "add_center",
            "step": "name",
            "data": {},
        }
        bot.edit_message_text(
            "🏫 <b>O'quv markazi qo'shish</b>\n\n📝 <b>1-qadam:</b> Markaz nomini kiriting:",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="HTML",
            reply_markup=admin_main_keyboard(),
        )
        bot.answer_callback_query(call.id)

    @bot.callback_query_handler(func=lambda c: c.data == "adm:del_center")
    def adm_del_center(call):
        if not is_admin(call.from_user.id):
            return
        admin_state[call.from_user.id] = {"action": "del_center"}
        bot.edit_message_text(
            "🗑️ <b>Markaz o'chirish</b>\n\n🔢 Markaz <b>ID</b> sini kiriting:",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="HTML",
            reply_markup=admin_main_keyboard(),
        )
        bot.answer_callback_query(call.id)

    # ─── UNIVERSITET ──────────────────────────────────────────────

    @bot.callback_query_handler(func=lambda c: c.data == "adm:add_uni")
    def adm_add_uni(call):
        if not is_admin(call.from_user.id):
            return
        admin_state[call.from_user.id] = {
            "action": "add_uni",
            "step": "name",
            "data": {},
        }
        bot.edit_message_text(
            "🎓 <b>Universitet qo'shish</b>\n\n📝 <b>1-qadam:</b> Nomini kiriting:",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="HTML",
            reply_markup=admin_main_keyboard(),
        )
        bot.answer_callback_query(call.id)

    @bot.callback_query_handler(func=lambda c: c.data == "adm:del_uni")
    def adm_del_uni(call):
        if not is_admin(call.from_user.id):
            return
        admin_state[call.from_user.id] = {"action": "del_uni"}
        bot.edit_message_text(
            "🗑️ <b>Universitet o'chirish</b>\n\n🔢 Universitet <b>ID</b> sini kiriting:",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="HTML",
            reply_markup=admin_main_keyboard(),
        )
        bot.answer_callback_query(call.id)

    # ─── BROADCAST: 1-qadam — Media ──────────────────────────────

    @bot.callback_query_handler(func=lambda c: c.data == "adm:broadcast")
    def adm_broadcast_start(call):
        if not is_admin(call.from_user.id):
            return
        admin_state[call.from_user.id] = {
            "action": "broadcast",
            "step": "media",
            "data": {},
        }
        bot.edit_message_text(
            "📢 <b>Reklama yuborish</b>\n\n"
            "🖼 <b>1-qadam:</b> Rasm, GIF yoki video yuboring.\n"
            "⏭️ Media kerak bo'lmasa — o'tkazib yuboring.",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="HTML",
            reply_markup=skip_media_keyboard(),
        )
        bot.answer_callback_query(call.id)

    # ─── BROADCAST: 3-qadam — Tugma qo'shish ────────────────────

    @bot.callback_query_handler(func=lambda c: c.data == "bcast:add_btn")
    def bcast_add_btn(call):
        if not is_admin(call.from_user.id):
            return
        uid = call.from_user.id
        state = admin_state.get(uid, {})
        state["step"] = "button_input"
        admin_state[uid] = state
        bot.send_message(
            call.message.chat.id,
            "🔗 <b>Inline tugma qo'shish</b>\n\n"
            "✏️ Format: <code>Tugma matni | https://havola.uz</code>\n\n"
            "📌 Misol: <code>Batafsil | https://t.me/Abituriyent_yordamchi</code>",
            parse_mode="HTML",
            reply_markup=skip_button_keyboard(),
        )
        bot.answer_callback_query(call.id)

    @bot.callback_query_handler(func=lambda c: c.data == "bcast:skip_btn")
    def bcast_skip_btn(call):
        if not is_admin(call.from_user.id):
            return
        uid = call.from_user.id
        state = admin_state.get(uid, {})
        state["data"]["buttons"] = []
        admin_state[uid] = state
        bot.answer_callback_query(call.id)
        _show_broadcast_preview(bot, call.message.chat.id, uid)

    # ─── BROADCAST: Tasdiqlash ────────────────────────────────────

    @bot.callback_query_handler(func=lambda c: c.data == "broadcast:confirm")
    def broadcast_confirm(call):
        if not is_admin(call.from_user.id):
            return
        uid = call.from_user.id
        state = admin_state.get(uid, {})
        data = state.get("data", {})

        preview_chat_id = data.get("preview_chat_id")
        preview_msg_id = data.get("preview_msg_id")

        if not preview_msg_id:
            bot.answer_callback_query(call.id, "⚠️ Xabar topilmadi!", show_alert=True)
            return

        # Inline tugmalar qayta tiklanadi (copy_message reply_markup'ni saqlamaydi)
        reply_markup = _build_buttons_markup(data.get("buttons", []))

        user_ids = get_all_user_ids()
        sent = []
        for target_uid in user_ids:
            try:
                # copy_message — "Forwarded from" CHIQMAYDI, tugma qayta beriladi
                msg = bot.copy_message(
                    target_uid, preview_chat_id, preview_msg_id,
                    reply_markup=reply_markup
                )
                sent.append((target_uid, msg.message_id))
            except Exception:
                pass

        if sent:
            save_broadcast(sent)

        try:
            bot.edit_message_text(
                f"✅ <b>Reklama yuborildi!</b>\n\n📤 {len(sent)} ta foydalanuvchiga yetkazildi.",
                call.message.chat.id,
                call.message.message_id,
                parse_mode="HTML",
                reply_markup=admin_main_keyboard(),
            )
        except Exception:
            bot.send_message(
                call.message.chat.id,
                f"✅ <b>Reklama yuborildi!</b>\n\n📤 {len(sent)} ta foydalanuvchiga yetkazildi.",
                parse_mode="HTML",
                reply_markup=admin_main_keyboard(),
            )
        admin_state.pop(uid, None)
        bot.answer_callback_query(call.id)

    @bot.callback_query_handler(func=lambda c: c.data == "broadcast:cancel")
    def broadcast_cancel(call):
        if not is_admin(call.from_user.id):
            return
        admin_state.pop(call.from_user.id, None)
        try:
            bot.edit_message_text(
                "❌ <b>Reklama bekor qilindi.</b>",
                call.message.chat.id,
                call.message.message_id,
                parse_mode="HTML",
                reply_markup=admin_main_keyboard(),
            )
        except Exception:
            bot.send_message(
                call.message.chat.id,
                "❌ <b>Reklama bekor qilindi.</b>",
                parse_mode="HTML",
                reply_markup=admin_main_keyboard(),
            )
        bot.answer_callback_query(call.id)

    # ─── REKLAMANI O'CHIRISH ──────────────────────────────────────

    @bot.callback_query_handler(func=lambda c: c.data == "adm:del_broadcast")
    def adm_del_broadcast(call):
        if not is_admin(call.from_user.id):
            return
        last = get_last_broadcast()
        if not last or not last.get("message_ids"):
            bot.answer_callback_query(
                call.id, "⚠️ O'chiriladigan reklama topilmadi!", show_alert=True
            )
            return
        deleted = 0
        for pair in last["message_ids"].split(","):
            try:
                uid_str, mid_str = pair.split(":")
                bot.delete_message(int(uid_str), int(mid_str))
                deleted += 1
            except Exception:
                pass
        bot.edit_message_text(
            f"🗑️ <b>Reklama o'chirildi!</b>\n\n✅ {deleted} ta xabar o'chirildi.",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="HTML",
            reply_markup=admin_main_keyboard(),
        )
        bot.answer_callback_query(call.id)

    # ─── ADMIN SELECT ─────────────────────────────────────────────

    @bot.callback_query_handler(func=lambda c: c.data.startswith("admsel:"))
    def adm_select(call):
        if not is_admin(call.from_user.id):
            return
        parts = call.data.split(":", 2)
        field, value = parts[1], parts[2]
        uid = call.from_user.id
        state = admin_state.get(uid, {})
        if "data" not in state:
            state["data"] = {}
        state["data"][field] = value
        admin_state[uid] = state
        bot.answer_callback_query(call.id, f"✅ {value} tanlandi")
        _advance_step(bot, call.message, uid)

    @bot.callback_query_handler(func=lambda c: c.data == "skip_media")
    def skip_media_cb(call):
        if not is_admin(call.from_user.id):
            return
        uid = call.from_user.id
        state = admin_state.get(uid, {})
        if "data" not in state:
            state["data"] = {}
        state["data"]["media_file_id"] = None
        state["data"]["media_type"] = None
        admin_state[uid] = state
        bot.answer_callback_query(call.id)

        if state.get("action") == "broadcast":
            state["step"] = "text"
            admin_state[uid] = state
            bot.send_message(
                call.message.chat.id,
                "✏️ <b>2-qadam:</b> Xabar matnini kiriting:\n\n"
                "💡 <i>Bold, italic, havolalar — hammasi saqlanadi.</i>",
                parse_mode="HTML",
                reply_markup=admin_main_keyboard(),
            )
        else:
            _advance_step(bot, call.message, uid)

    @bot.callback_query_handler(func=lambda c: c.data == "skip_button")
    def skip_button_cb(call):
        if not is_admin(call.from_user.id):
            return
        uid = call.from_user.id
        state = admin_state.get(uid, {})
        if "data" not in state:
            state["data"] = {}
        state["data"]["buttons"] = []
        admin_state[uid] = state
        bot.answer_callback_query(call.id)
        _advance_step(bot, call.message, uid)

    # ─── O'RIN TANLASH (QO'SHISH VAQTIDA) ────────────────────────

    @bot.callback_query_handler(func=lambda c: c.data.startswith("admpos:"))
    def admpos_cb(call):
        if not is_admin(call.from_user.id):
            return
        uid = call.from_user.id
        state = admin_state.get(uid, {})
        action = state.get("action")
        data = state.get("data", {})
        chat_id = call.message.chat.id
        try:
            pos = int(call.data.split(":")[1])
        except (IndexError, ValueError):
            bot.answer_callback_query(call.id)
            return

        if action == "add_teacher":
            row_id = add_teacher(
                name=data.get("name", ""),
                city=data.get("city", ""),
                subject=data.get("subject", ""),
                teach_type=data.get("teach_type", ""),
                description=data.get("description", ""),
                media_file_id=data.get("media_file_id"),
                media_type=data.get("media_type"),
                buttons=data.get("buttons", []),
            )
            set_item_position("teachers", row_id, pos)
            bot.send_message(
                chat_id,
                f"✅ <b>Ustoz qo'shildi!</b> 👨‍🏫\n\n"
                f"🆔 ID: <code>{row_id}</code>\n"
                f"👤 Ism: <b>{data.get('name')}</b>\n"
                f"📚 Fan: <b>{data.get('subject')}</b>\n"
                f"🏙️ Shahar: <b>{data.get('city')}</b>\n"
                f"📍 O'rin: <b>{pos}</b>",
                parse_mode="HTML",
                reply_markup=admin_main_keyboard(),
            )
        elif action == "add_center":
            row_id = add_center(
                name=data.get("name", ""),
                city=data.get("city", ""),
                description=data.get("description", ""),
                media_file_id=data.get("media_file_id"),
                media_type=data.get("media_type"),
                buttons=data.get("buttons", []),
            )
            set_item_position("centers", row_id, pos)
            bot.send_message(
                chat_id,
                f"✅ <b>Markaz qo'shildi!</b> 🏫\n\n"
                f"🆔 ID: <code>{row_id}</code>\n"
                f"🏢 Nom: <b>{data.get('name')}</b>\n"
                f"🏙️ Shahar: <b>{data.get('city')}</b>\n"
                f"📍 O'rin: <b>{pos}</b>",
                parse_mode="HTML",
                reply_markup=admin_main_keyboard(),
            )
        elif action == "add_uni":
            row_id = add_university(
                name=data.get("name", ""),
                uni_type=data.get("uni_type", ""),
                description=data.get("description", ""),
                media_file_id=data.get("media_file_id"),
                media_type=data.get("media_type"),
                buttons=data.get("buttons", []),
            )
            set_item_position("universities", row_id, pos)
            bot.send_message(
                chat_id,
                f"✅ <b>Universitet qo'shildi!</b> 🎓\n\n"
                f"🆔 ID: <code>{row_id}</code>\n"
                f"🏛️ Nom: <b>{data.get('name')}</b>\n"
                f"📋 Tur: <b>{data.get('uni_type')}</b>\n"
                f"📍 O'rin: <b>{pos}</b>",
                parse_mode="HTML",
                reply_markup=admin_main_keyboard(),
            )
        else:
            bot.answer_callback_query(call.id)
            return

        admin_state.pop(uid, None)
        bot.answer_callback_query(call.id, f"✅ {pos}-o'ringa joylashtirildi!")

    # ─── TARTIB TAHRIRLASH ────────────────────────────────────────

    @bot.callback_query_handler(func=lambda c: c.data == "adm:edit_positions")
    def adm_edit_positions(call):
        if not is_admin(call.from_user.id):
            return
        admin_state.pop(call.from_user.id, None)
        try:
            bot.edit_message_text(
                "📋 <b>Tartibni tahrirlash</b>\n\nQaysi bo'limni o'zgartirmoqchisiz?",
                call.message.chat.id,
                call.message.message_id,
                parse_mode="HTML",
                reply_markup=edit_positions_menu_keyboard(),
            )
        except Exception:
            bot.send_message(
                call.message.chat.id,
                "📋 <b>Tartibni tahrirlash</b>\n\nQaysi bo'limni o'zgartirmoqchisiz?",
                parse_mode="HTML",
                reply_markup=edit_positions_menu_keyboard(),
            )
        bot.answer_callback_query(call.id)

    @bot.callback_query_handler(func=lambda c: c.data.startswith("adm:editpos:"))
    def adm_editpos(call):
        if not is_admin(call.from_user.id):
            return
        table = call.data.split(":")[2]
        labels = {"teachers": "Ustozlar", "centers": "Markazlar", "universities": "Universitetlar"}
        label = labels.get(table, table)
        items = get_positions_list(table)
        if not items:
            bot.answer_callback_query(call.id, f"⚠️ {label} bo'sh!", show_alert=True)
            return
        lines = "\n".join(f"  {i}. {item['name']}" for i, item in enumerate(items, 1))
        text = f"📋 <b>{label} tartibi:</b>\n\n{lines}\n\n✏️ Qaysi yozuvni ko'chirmoqchisiz?"
        try:
            bot.edit_message_text(
                text, call.message.chat.id, call.message.message_id,
                parse_mode="HTML", reply_markup=items_list_keyboard(items, table),
            )
        except Exception:
            bot.send_message(
                call.message.chat.id, text,
                parse_mode="HTML", reply_markup=items_list_keyboard(items, table),
            )
        bot.answer_callback_query(call.id)

    @bot.callback_query_handler(func=lambda c: c.data.startswith("adm:moveitem:"))
    def adm_moveitem(call):
        if not is_admin(call.from_user.id):
            return
        parts = call.data.split(":")
        table = parts[2]
        item_id = int(parts[3])
        labels = {"teachers": "Ustozlar", "centers": "Markazlar", "universities": "Universitetlar"}
        label = labels.get(table, table)
        items = get_positions_list(table)
        current_item = next((x for x in items if x["id"] == item_id), None)
        if not current_item:
            bot.answer_callback_query(call.id, "⚠️ Element topilmadi!", show_alert=True)
            return
        current_pos = next((i for i, x in enumerate(items, 1) if x["id"] == item_id), 0)
        lines = "\n".join(
            f"  ➡️ {x['name']}" if x["id"] == item_id else f"  {i}. {x['name']}"
            for i, x in enumerate(items, 1)
        )
        text = (
            f"📋 <b>{label} tartibi:</b>\n\n{lines}\n\n"
            f"✏️ <b>{current_item['name']}</b> hozir <b>{current_pos}-o'rinda</b>.\n"
            f"Yangi o'rinni tanlang (● = hozirgi o'rin):"
        )
        try:
            bot.edit_message_text(
                text, call.message.chat.id, call.message.message_id,
                parse_mode="HTML", reply_markup=move_item_keyboard(items, table, item_id),
            )
        except Exception:
            bot.send_message(
                call.message.chat.id, text,
                parse_mode="HTML", reply_markup=move_item_keyboard(items, table, item_id),
            )
        bot.answer_callback_query(call.id)

    @bot.callback_query_handler(func=lambda c: c.data.startswith("adm:setpos:"))
    def adm_setpos(call):
        if not is_admin(call.from_user.id):
            return
        parts = call.data.split(":")
        table = parts[2]
        item_id = int(parts[3])
        new_pos = int(parts[4])
        labels = {"teachers": "Ustozlar", "centers": "Markazlar", "universities": "Universitetlar"}
        label = labels.get(table, table)
        ok = set_item_position(table, item_id, new_pos)
        if not ok:
            bot.answer_callback_query(call.id, "⚠️ Xatolik yuz berdi!", show_alert=True)
            return
        items = get_positions_list(table)
        lines = "\n".join(f"  {i}. {item['name']}" for i, item in enumerate(items, 1))
        text = (
            f"✅ <b>O'rin o'zgartirildi!</b>\n\n"
            f"📋 <b>{label} yangi tartibi:</b>\n\n{lines}\n\n"
            f"✏️ Yana ko'chirmoqchisiz?"
        )
        try:
            bot.edit_message_text(
                text, call.message.chat.id, call.message.message_id,
                parse_mode="HTML", reply_markup=items_list_keyboard(items, table),
            )
        except Exception:
            bot.send_message(
                call.message.chat.id, text,
                parse_mode="HTML", reply_markup=items_list_keyboard(items, table),
            )
        bot.answer_callback_query(call.id, f"✅ {new_pos}-o'ringa ko'chirildi!")

    # ─── MATN XABARLARINI QAYTA ISHLASH ─────────────────────────

    @bot.message_handler(
        func=lambda m: m.from_user.id in admin_state,
        content_types=["text", "photo", "video", "animation", "document"],
    )
    def admin_text_handler(message):
        uid = message.from_user.id
        if not is_admin(uid):
            return
        state = admin_state.get(uid, {})
        action = state.get("action")
        step = state.get("step")

        # ── BROADCAST ─────────────────────────────────────────────
        if action == "broadcast":
            if step == "media":
                if message.content_type == "photo":
                    state["data"]["media_file_id"] = message.photo[-1].file_id
                    state["data"]["media_type"] = "photo"
                elif message.content_type == "video":
                    state["data"]["media_file_id"] = message.video.file_id
                    state["data"]["media_type"] = "video"
                elif message.content_type == "animation":
                    state["data"]["media_file_id"] = message.animation.file_id
                    state["data"]["media_type"] = "animation"
                else:
                    bot.send_message(
                        message.chat.id,
                        "❌ Faqat rasm 🖼, video 🎬 yoki GIF yuboring!",
                        reply_markup=skip_media_keyboard(),
                    )
                    return
                state["step"] = "text"
                admin_state[uid] = state
                bot.send_message(
                    message.chat.id,
                    "✏️ <b>2-qadam:</b> Xabar matnini kiriting:\n\n"
                    "💡 <i>Bold, italic, havolalar — hammasi saqlanadi.</i>",
                    parse_mode="HTML",
                    reply_markup=admin_main_keyboard(),
                )
                return

            if step == "text":
                if message.content_type != "text":
                    bot.send_message(message.chat.id, "✏️ Iltimos matn kiriting!")
                    return
                # html_text — admin yozgan formatlash (bold, italic, havolalar) saqlanadi
                state["data"]["text"] = message.html_text or message.text or ""
                state["step"] = "button"
                admin_state[uid] = state
                bot.send_message(
                    message.chat.id,
                    "🔗 <b>3-qadam:</b> Inline tugma qo'shasizmi?",
                    parse_mode="HTML",
                    reply_markup=broadcast_btn_step_keyboard(),
                )
                return

            if step == "button_input":
                if message.content_type != "text":
                    bot.send_message(
                        message.chat.id,
                        "❌ Format: <code>Matn | URL</code>",
                        parse_mode="HTML",
                        reply_markup=skip_button_keyboard(),
                    )
                    return
                txt = message.text.strip()
                # Har bir qatorda bitta tugma: "Matn | URL" — eng ko'pi 3 ta
                buttons = []
                for line in txt.splitlines():
                    line = line.strip()
                    if "|" in line:
                        p = line.split("|", 1)
                        bt, bu = p[0].strip(), p[1].strip()
                        if bt and bu:
                            buttons.append((bt, bu))
                    if len(buttons) >= 3:
                        break

                if not buttons:
                    bot.send_message(
                        message.chat.id,
                        "❌ Format noto'g'ri!\n\n"
                        "Har qatorda <b>bitta</b> tugma:\n"
                        "<code>Tugma matni | https://havola.uz</code>\n\n"
                        "📌 Misol (3 ta tugma):\n"
                        "<code>Kanal | https://t.me/Abituriyent_yordamchi\n"
                        "Sayt | https://example.uz\n"
                        "Aloqa | https://t.me/Adm1nnn_1_0</code>",
                        parse_mode="HTML",
                        reply_markup=skip_button_keyboard(),
                    )
                    return

                state["data"]["buttons"] = buttons
                admin_state[uid] = state
                bot.send_message(
                    message.chat.id,
                    f"✅ <b>{len(buttons)} ta</b> tugma qo'shildi!",
                    parse_mode="HTML",
                )
                _show_broadcast_preview(bot, message.chat.id, uid)
                return
            return

        # ── Shahar ────────────────────────────────────────────────
        if action == "add_city":
            if not message.text:
                return
            name = message.text.strip()
            ok = add_city(name)
            bot.send_message(
                message.chat.id,
                f"✅ <b>'{name}'</b> qo'shildi! 🏙️"
                if ok
                else f"⚠️ <b>'{name}'</b> allaqachon mavjud!",
                parse_mode="HTML",
                reply_markup=admin_main_keyboard(),
            )
            admin_state.pop(uid, None)
            return

        # ── Fan ───────────────────────────────────────────────────
        if action == "add_subject":
            if not message.text:
                return
            name = message.text.strip()
            ok = add_subject(name)
            bot.send_message(
                message.chat.id,
                f"✅ <b>'{name}'</b> qo'shildi! 📚"
                if ok
                else f"⚠️ <b>'{name}'</b> allaqachon mavjud!",
                parse_mode="HTML",
                reply_markup=admin_main_keyboard(),
            )
            admin_state.pop(uid, None)
            return

        # ── O'chirish ─────────────────────────────────────────────
        if action == "del_teacher":
            _delete_by_id(bot, message, uid, delete_teacher, "Ustoz", "👨‍🏫")
            return
        if action == "del_center":
            _delete_by_id(bot, message, uid, delete_center, "Markaz", "🏫")
            return
        if action == "del_uni":
            _delete_by_id(bot, message, uid, delete_university, "Universitet", "🎓")
            return

        # ── Qadamli qo'shish ──────────────────────────────────────
        if action in ("add_teacher", "add_center", "add_uni"):
            if "data" not in state:
                state["data"] = {}

            # Media qabul
            if step == "media":
                if message.content_type == "photo":
                    state["data"]["media_file_id"] = message.photo[-1].file_id
                    state["data"]["media_type"] = "photo"
                elif message.content_type == "video":
                    state["data"]["media_file_id"] = message.video.file_id
                    state["data"]["media_type"] = "video"
                elif message.content_type == "animation":
                    state["data"]["media_file_id"] = message.animation.file_id
                    state["data"]["media_type"] = "animation"
                else:
                    bot.send_message(
                        message.chat.id,
                        "❌ Faqat rasm 🖼, video 🎬 yoki GIF yuboring!",
                        reply_markup=skip_media_keyboard(),
                    )
                    return
                admin_state[uid] = state
                _advance_step(bot, message, uid)
                return

            # Inline tugmalar (eng ko'pi 3 ta — har qatorda bittadan)
            if step == "button" and message.content_type == "text":
                txt = message.text.strip()
                buttons = []
                for line in txt.splitlines():
                    line = line.strip()
                    if "|" in line:
                        p = line.split("|", 1)
                        bt, bu = p[0].strip(), p[1].strip()
                        if bt and bu:
                            buttons.append((bt, bu))
                    if len(buttons) >= 3:
                        break
                if not buttons:
                    bot.send_message(
                        message.chat.id,
                        "❌ Format noto'g'ri!\n\n"
                        "Har qatorda <b>bitta</b> tugma (eng ko'pi 3 ta):\n"
                        "<code>Matn | https://URL</code>\n\n"
                        "📌 Misol:\n"
                        "<code>Batafsil | https://t.me/example\n"
                        "Sayt | https://example.uz\n"
                        "Aloqa | https://t.me/Adm1nnn_1_0</code>",
                        parse_mode="HTML",
                        reply_markup=skip_button_keyboard(),
                    )
                    return
                state["data"]["buttons"] = buttons
                admin_state[uid] = state
                _advance_step(bot, message, uid)
                return

            # Matn
            if message.content_type == "text":
                # Tavsifda formatlash (bold, italic, havolalar) saqlansin
                if step == "description":
                    state["data"][step] = message.html_text or message.text or ""
                else:
                    state["data"][step] = message.text.strip()
                admin_state[uid] = state
                _advance_step(bot, message, uid)
                return

        admin_state[uid] = state


# ─── YORDAMCHI FUNKSIYALAR ────────────────────────────────────────


def _build_position_text(items, label):
    """O'rin tanlash uchun matn: hozirgi tartib + so'rov."""
    if not items:
        return (
            f"📋 Hozir <b>{label}</b> mavjud emas.\n\n"
            f"📍 Yangi yozuv qaysi o'ringa joylashsin?"
        )
    lines = "\n".join(f"  {i}. {item['name']}" for i, item in enumerate(items, 1))
    return (
        f"📋 <b>Hozirgi tartib ({label}):</b>\n{lines}\n\n"
        f"📍 Yangi yozuv qaysi o'ringa joylashsin?"
    )


def _build_buttons_markup(buttons):
    """[(text, url), ...] dan InlineKeyboardMarkup yasaydi (har bir tugma alohida qatorda)."""
    if not buttons:
        return None
    kb = types.InlineKeyboardMarkup()
    for bt, bu in buttons[:3]:
        kb.add(types.InlineKeyboardButton(bt, url=bu))
    return kb


def _send_xlsx(bot, call, title, headers, data, filename):
    """Excel fayl yaratib admin ga yuboradi."""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="2E7D32")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for row in data:
        ws.append(row)

    # Ustun kengligini sozlash
    for i, h in enumerate(headers, 1):
        max_len = len(str(h))
        for r in data:
            v = r[i - 1] if i - 1 < len(r) else ""
            max_len = max(max_len, len(str(v)) if v else 0)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = filename

    if not data:
        bot.answer_callback_query(call.id, f"⚠️ {title} hali bo'sh!", show_alert=True)
        return

    bot.send_document(
        call.message.chat.id,
        buf,
        visible_file_name=filename,
        caption=f"📊 <b>{title}</b>\n📦 Jami: <b>{len(data)}</b> ta yozuv",
        parse_mode="HTML"
    )
    bot.answer_callback_query(call.id, f"✅ {title} fayli yuborildi!")


def _delete_by_id(bot, message, uid, delete_fn, label, emoji):
    try:
        item_id = int(message.text.strip())
        ok = delete_fn(item_id)
        if ok:
            bot.send_message(
                message.chat.id,
                f"✅ {emoji} <b>{label} #{item_id}</b> o'chirildi!",
                parse_mode="HTML",
                reply_markup=admin_main_keyboard(),
            )
        else:
            bot.send_message(
                message.chat.id,
                f"❌ <b>ID #{item_id}</b> topilmadi!",
                parse_mode="HTML",
                reply_markup=admin_main_keyboard(),
            )
    except (ValueError, AttributeError):
        bot.send_message(
            message.chat.id,
            "❌ Faqat raqam kiriting!",
            reply_markup=admin_main_keyboard(),
        )
    admin_state.pop(uid, None)


def _show_broadcast_preview(bot, chat_id, uid):
    """Preview yuboradi va copy_message uchun preview_msg_id saqlaydi."""
    state = admin_state.get(uid, {})
    data = state.get("data", {})
    text = data.get("text", "")
    media_file_id = data.get("media_file_id")
    media_type = data.get("media_type")
    reply_markup = _build_buttons_markup(data.get("buttons", []))

    bot.send_message(
        chat_id, "👁 <b>Ko'rinish (Preview):</b>\n━━━━━━━━━━━━━━━━━", parse_mode="HTML"
    )

    # Preview xabarini yuborish va ID sini saqlash
    preview_msg = _send_direct(
        bot, chat_id, text, media_file_id, media_type, reply_markup
    )

    bot.send_message(
        chat_id,
        "━━━━━━━━━━━━━━━━━\n✅ Yuborishni tasdiqlaysizmi?",
        reply_markup=confirm_broadcast_keyboard(),
    )

    if preview_msg:
        state["data"]["preview_msg_id"] = preview_msg.message_id
        state["data"]["preview_chat_id"] = chat_id
    state["step"] = "confirm"
    admin_state[uid] = state


def _send_direct(bot, chat_id, text, media_file_id, media_type, reply_markup=None):
    """To'g'ridan-to'g'ri yuboradi — admin ismi CHIQMAYDI."""
    try:
        if media_file_id and media_type == "photo":
            return bot.send_photo(
                chat_id,
                media_file_id,
                caption=text,
                parse_mode="HTML",
                reply_markup=reply_markup,
            )
        elif media_file_id and media_type == "video":
            return bot.send_video(
                chat_id,
                media_file_id,
                caption=text,
                parse_mode="HTML",
                reply_markup=reply_markup,
            )
        elif media_file_id and media_type == "animation":
            return bot.send_animation(
                chat_id,
                media_file_id,
                caption=text,
                parse_mode="HTML",
                reply_markup=reply_markup,
            )
        else:
            return bot.send_message(
                chat_id, text, parse_mode="HTML", reply_markup=reply_markup
            )
    except Exception:
        return None


# ─── QADAMLI QO'SHISH ─────────────────────────────────────────────


def _advance_step(bot, message, uid):
    state = admin_state.get(uid, {})
    action = state.get("action")
    data = state.get("data", {})
    chat_id = message.chat.id

    # ── O'QITUVCHI ────────────────────────────────────────────────
    if action == "add_teacher":
        step = state.get("step")

        if step == "name":
            state["step"] = "city"
            bot.send_message(
                chat_id,
                "🏙️ <b>2-qadam:</b> Shaharni tanlang:",
                parse_mode="HTML",
                reply_markup=admin_select_keyboard(get_cities(), "city"),
            )
        elif step == "city":
            state["step"] = "subject"
            bot.send_message(
                chat_id,
                "📚 <b>3-qadam:</b> Fanni tanlang:",
                parse_mode="HTML",
                reply_markup=admin_select_keyboard(get_subjects(), "subject"),
            )
        elif step == "subject":
            state["step"] = "teach_type"
            bot.send_message(
                chat_id,
                "🖥️ <b>4-qadam:</b> Dars turini tanlang:",
                parse_mode="HTML",
                reply_markup=admin_select_keyboard(["Online", "Offline"], "teach_type"),
            )
        elif step == "teach_type":
            state["step"] = "media"
            bot.send_message(
                chat_id,
                "🖼 <b>5-qadam:</b> Rasm, video yoki GIF yuboring:",
                parse_mode="HTML",
                reply_markup=skip_media_keyboard(),
            )
        elif step == "media":
            state["step"] = "description"
            bot.send_message(
                chat_id,
                "📝 <b>6-qadam:</b> Tavsif kiriting:",
                parse_mode="HTML",
                reply_markup=admin_main_keyboard(),
            )
        elif step == "description":
            state["step"] = "button"
            bot.send_message(
                chat_id,
                "🔗 <b>7-qadam:</b> Inline tugmalar (eng ko'pi <b>3 ta</b>, ixtiyoriy):\n\n"
                "Har qatorda bittadan: <code>Matn | https://URL</code>\n\n"
                "📌 Misol:\n"
                "<code>Batafsil | https://t.me/example\n"
                "Aloqa | https://t.me/Adm1nnn_1_0</code>",
                parse_mode="HTML",
                reply_markup=skip_button_keyboard(),
            )
        elif step == "button":
            items = get_positions_list("teachers")
            state["step"] = "position_wait"
            admin_state[uid] = state
            bot.send_message(
                chat_id,
                _build_position_text(items, "ustozlar"),
                parse_mode="HTML",
                reply_markup=position_select_keyboard(items),
            )
            return

        admin_state[uid] = state

    # ── MARKAZ ────────────────────────────────────────────────────
    elif action == "add_center":
        step = state.get("step")

        if step == "name":
            state["step"] = "city"
            bot.send_message(
                chat_id,
                "🏙️ <b>2-qadam:</b> Shaharni tanlang:",
                parse_mode="HTML",
                reply_markup=admin_select_keyboard(get_cities(), "city"),
            )
        elif step == "city":
            state["step"] = "media"
            bot.send_message(
                chat_id,
                "🖼 <b>3-qadam:</b> Rasm, video yoki GIF yuboring:",
                parse_mode="HTML",
                reply_markup=skip_media_keyboard(),
            )
        elif step == "media":
            state["step"] = "description"
            bot.send_message(
                chat_id,
                "📝 <b>4-qadam:</b> Tavsif kiriting:",
                parse_mode="HTML",
                reply_markup=admin_main_keyboard(),
            )
        elif step == "description":
            state["step"] = "button"
            bot.send_message(
                chat_id,
                "🔗 <b>5-qadam:</b> Inline tugmalar (eng ko'pi <b>3 ta</b>, ixtiyoriy):\n\n"
                "Har qatorda bittadan: <code>Matn | https://URL</code>\n\n"
                "📌 Misol:\n"
                "<code>Batafsil | https://t.me/example\n"
                "Aloqa | https://t.me/Adm1nnn_1_0</code>",
                parse_mode="HTML",
                reply_markup=skip_button_keyboard(),
            )
        elif step == "button":
            items = get_positions_list("centers")
            state["step"] = "position_wait"
            admin_state[uid] = state
            bot.send_message(
                chat_id,
                _build_position_text(items, "markazlar"),
                parse_mode="HTML",
                reply_markup=position_select_keyboard(items),
            )
            return

        admin_state[uid] = state

    # ── UNIVERSITET ───────────────────────────────────────────────
    elif action == "add_uni":
        step = state.get("step")

        if step == "name":
            state["step"] = "uni_type"
            bot.send_message(
                chat_id,
                "🏛️ <b>2-qadam:</b> Turini tanlang:",
                parse_mode="HTML",
                reply_markup=admin_select_keyboard(["Davlat", "Nodavlat"], "uni_type"),
            )
        elif step == "uni_type":
            state["step"] = "media"
            bot.send_message(
                chat_id,
                "🖼 <b>3-qadam:</b> Rasm, video yoki GIF yuboring:",
                parse_mode="HTML",
                reply_markup=skip_media_keyboard(),
            )
        elif step == "media":
            state["step"] = "description"
            bot.send_message(
                chat_id,
                "📝 <b>4-qadam:</b> Tavsif kiriting:",
                parse_mode="HTML",
                reply_markup=admin_main_keyboard(),
            )
        elif step == "description":
            state["step"] = "button"
            bot.send_message(
                chat_id,
                "🔗 <b>5-qadam:</b> Inline tugmalar (eng ko'pi <b>3 ta</b>, ixtiyoriy):\n\n"
                "Har qatorda bittadan: <code>Matn | https://URL</code>\n\n"
                "📌 Misol:\n"
                "<code>Batafsil | https://t.me/example\n"
                "Aloqa | https://t.me/Adm1nnn_1_0</code>",
                parse_mode="HTML",
                reply_markup=skip_button_keyboard(),
            )
        elif step == "button":
            items = get_positions_list("universities")
            state["step"] = "position_wait"
            admin_state[uid] = state
            bot.send_message(
                chat_id,
                _build_position_text(items, "universitetlar"),
                parse_mode="HTML",
                reply_markup=position_select_keyboard(items),
            )
            return

        admin_state[uid] = state
